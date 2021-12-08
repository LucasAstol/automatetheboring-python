#! python3
# sendDuesReminders.py - sends reminder emails to members that have due payments

import openpyxl, smtplib, sys, ezgmail

wb = openpyxl.load_workbook(r'/scratch/home/lastolfi/Education/automatetheboring-python/automate_online-materials/duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

unpaidMembers = {}
for r in range(2, sheet.max_row +1):
    if(sheet.cell(row=r, column=lastCol).value != 'paid'):
        name = sheet.cell(row=r, column=1).value
        theEmail = sheet.cell(row=r, column=2)
        unpaidMembers[name] = theEmail

for name,email in unpaidMembers.items():
    dueMessage = ezgmail._createMessage('',email, latestMonth + ' - Due payment', 'Dear ' + name + '\nPlease pay as soon as possible.')
    gmailConn = ezgmail._sendMessage(dueMessage)
