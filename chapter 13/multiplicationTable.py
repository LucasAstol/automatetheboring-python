#! python3
# multiplicationTable.py - creates a multiplication table of N*N being N a parameter passed by command line

import openpyxl, sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active

dimension = int(sys.argv[1])
dimensionRange = dimension+2

for i in range(1, dimensionRange):
        if i == 1:
            for x in range(2, dimensionRange):
                fontObj = Font(bold=True)
                
                horizontalCell = sheet.cell(row=i, column=x)
                horizontalCell.value = x-1
                horizontalCell.font = fontObj

                verticalCell = sheet.cell(row=x, column=i)
                verticalCell.value = x-1
                verticalCell.font = fontObj 
        else:
            for x in range(2, dimensionRange):
                coord1 = 'A' + str(i)
                coord2 = get_column_letter(x) + str(1)
                sheet.cell(row=i, column=x).value = f'={coord1}*{coord2}'

wb.save('C:\\Lucas\\Educacion\\Automate the boring stuff\\chapter 13\\multiplicationTable.xlsx')

