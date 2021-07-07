#! pyhton3
# changePhones.py - find phones and format them in an unique format

import openpyxl, re

numbers = ('1','2','3','4','5','6','7','8','9','0')
print('Opening workbook...')
wb = openpyxl.load_workbook('C:\\Lucas\\Educacion\\Automate the boring stuff\\chapter 13\\phones.xlsx')
sheet = wb.active

for rowNum in range(2, sheet.max_row + 1):
    phoneNum = sheet.cell(row=rowNum, column=2).value
    newPhoneNum = ''
    for theChar in phoneNum:
        if theChar in numbers:
            newPhoneNum = newPhoneNum + theChar

    print('Saving phone...')
    sheet.cell(row=rowNum, column=2).value = newPhoneNum

wb.save('C:\\Lucas\\Educacion\\Automate the boring stuff\\chapter 13\\phonesCopy.xlsx')
print('Phones copy saved')