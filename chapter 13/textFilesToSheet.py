#! python3
# textFilesToSheet.py - copies each textfile line to a cell. One column per file.

import openpyxl, os

newWb = openpyxl.Workbook()
sheet = newWb.active
thePath = 'C:\\Lucas\\Educacion\\Automate the boring stuff\\chapter 13\\textFiles'
textFiles = os.listdir(thePath)

for i in range(len(textFiles)):
    filePath = os.path.join(thePath, textFiles[i])
    theFile = open(filePath)
    theLines = theFile.readlines()
    for x in range(len(theLines)):
        if x+1 == 1:
            sheet.cell(row=1, column=i+1).value = os.path.basename(filePath)
        sheet.cell(row=x+2, column=i+1).value = theLines[x]

newWb.save('C:\\Lucas\\Educacion\\Automate the boring stuff\\chapter 13\\textFiles.xlsx')