#! python3 
# changePrices.py

import openpyxl

wb = openpyxl.load_workbook('C:\\Lucas\\Educacion\\Automate the boring stuff\\automate_online-materials\\produceSales.xlsx')
sheet = wb['Sheet']

PRICE_UPDATES = {'Garlic': 1.89, 'Celery': 3.2, 'Lemon':4.21}

for rowNum in range(2, sheet.max_row + 1):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('newPricesProduceSales.xlsx')