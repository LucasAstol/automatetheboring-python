#! python3
# blankRowInserter.py - inserts a specific number of rows after an specific row

import openpyxl, sys

wb = openpyxl.load_workbook(sys.argv[3])
sheet = wb.active
startRow = int(sys.argv[1])
rows = int(sys.argv[2]) 

for i in range(sheet.max_row, -1, -1):
    for x in range(1, sheet.max_column + 1):
        if i >= int(sys.argv[1]):
            sheet.cell(row=i+rows, column=x).value = sheet.cell(row=i, column=x).value
            sheet.cell(row=i, column=x).value = ''
            
wb.save(sys.argv[3])        