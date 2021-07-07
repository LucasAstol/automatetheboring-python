#! python3
# getAllProductSales.py - returns a new spreadsheet with all the sales of an specific product 

import openpyxl, sys, os
from openpyxl.styles import Font

product = sys.argv[1] #'Potatoes'

wb = openpyxl.load_workbook('C:\\Lucas\Educacion\\Automate the boring stuff\\automate_online-materials\\produceSales.xlsx')
newWb = openpyxl.Workbook()

sheet = wb.active

rowNewWb = 0
for i in range(2, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == product:
        rowNewWb += 1 
        for x in range(1, sheet.max_column + 1):
            if x != 4:
                newWb.active.cell(row=rowNewWb, column=x).value = sheet.cell(row=i, column=x).value
            else:
                roundFormula = f'=ROUND(B{rowNewWb}*C{rowNewWb},2)'
                newWb.active.cell(row=rowNewWb, column=x).value = roundFormula
        newWb.active.cell(row=rowNewWb, column=1).font = Font(size=12, bold=True)

newWb.save(os.path.join('C:\\Lucas\Educacion\\Automate the boring stuff\\chapter 13', product + 'Sales.xlsx'))