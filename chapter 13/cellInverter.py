#! python3
# cellInverter.py - inverts the row and columns of the cell coyping the value to the new cell

import openpyxl

wb = openpyxl.load_workbook('C:\\Lucas\\Educacion\\Automate the boring stuff\\chapter 13\\examples_copy.xlsx')
sheet = wb.active
newWb = openpyxl.Workbook()

for i in range(1, sheet.max_column + 1):
    for x in range(1, sheet.max_row + 1):
        newWb.active.cell(row=i, column=x).value = sheet.cell(row=x, column=i).value

newWb.save('C:\\Lucas\\Educacion\\Automate the boring stuff\\chapter 13\\inverted_examples_copy.xlsx')