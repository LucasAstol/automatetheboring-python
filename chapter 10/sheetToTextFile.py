#! python3
# sheetToTextFile.py - takes each of the cells in a column and writes it to a file per column.

import openpyxl, os

wb = openpyxl.load_workbook('C:\\Lucas\\Educacion\\Automate the boring stuff\\chapter 13\\textFiles.xlsx')
sheet = wb.active

for i in range(1, sheet.max_column + 1):
    filename = sheet.cell(row=1, column=i).value
    theFile = open(os.path.join('C:\\Lucas\\Educacion\\Automate the boring stuff\\chapter 13\\fromSheet', filename + '.txt'), 'w')
    for x in range(2, sheet.max_row + 1):
        cellContent = sheet.cell(row=x, column=i).value
        if cellContent == None:
            theFile.write('')
        else:
            theFile.write(sheet.cell(row=x, column=i).value)

    theFile.close()