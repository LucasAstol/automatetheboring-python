#! python3 
# excelToCSV.py - converts excel files in a given folder to csv files.

import openpyxl, csv, os, sys
from pathlib import Path

theOrigin = Path(sys.argv[1])
destination = Path(sys.argv[2])


if len(sys.argv) == 3:
    if theOrigin.is_dir() and theOrigin.exists():
        for excelFile in list(theOrigin.glob("*.xlsx")):
            excelFilePath = Path(excelFile)
            theExcel = openpyxl.load_workbook(str(excelFilePath))

            for sheetname in theExcel.sheetnames:
                excelSheet = theExcel[sheetname]
                newCsvFile = open(Path(destination / (excelFilePath.stem + "-" + sheetname + ".csv")), 'w', newline = '')
                csvWriter = csv.writer(newCsvFile)

                for theRow in range(1, excelSheet.max_row + 1):
                    rowToWrite = []
                    for theCol in range(1, excelSheet.max_column + 1):
                        rowToWrite.append(excelSheet.cell(row=theRow, column=theCol).value)
                    csvWriter.writerow(rowToWrite)
                newCsvFile.close()
